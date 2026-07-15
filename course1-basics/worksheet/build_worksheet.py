from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/andy/Work/ad-edu/harness-ad-academy/course1-basics/worksheet/C1_실습_워크시트.xlsx"

NAVY = "1E293B"; CORAL = "FF6B4A"; TEAL = "2DD4BF"; TEALD = "0D9488"
PAPER = "F8FAFC"; MIST = "E2E8F0"; SLATE = "64748B"; INK = "1E293B"
AMBER_BG = "FEF3C7"; CORAL_BG = "FFF1EC"; TEAL_BG = "ECFDF5"; MISTBG = "F1F5F9"
FONT = "Pretendard"

thin = Side(style="thin", color=MIST)
med = Side(style="medium", color=NAVY)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=10, b=False, color=INK, it=False):
    return Font(name=FONT, size=sz, bold=b, color=color, italic=it)
def fill(c):
    return PatternFill("solid", fgColor=c)
def setc(ws, ref, val, font=None, bg=None, align=None, border=None, wrap=False):
    c = ws[ref]; c.value = val
    if font: c.font = font
    if bg: c.fill = fill(bg)
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "left", vertical="center", wrap_text=wrap)
    else:
        c.alignment = Alignment(vertical="center")
    if border: c.border = border
    return c

def title_block(ws, subtitle, start_col_span):
    setc(ws, "B2", "HARNESS AD ACADEMY", F(9, True, TEALD))
    setc(ws, "B3", "과정 1 · 기초 — 실습 워크시트", F(10, False, SLATE))
    setc(ws, "B4", subtitle, F(20, True, NAVY))
    # accent bar row 5
    for col in range(2, start_col_span + 1):
        ws.cell(row=5, column=col).fill = fill(CORAL)
    ws.row_dimensions[5].height = 4

wb = Workbook()

# ============================================================
# SHEET 1 — 안내
# ============================================================
ws = wb.active; ws.title = "안내"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2.5
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 78
title_block(ws, "실습 워크시트 · 사용 안내", 3)

setc(ws, "B7", "이 워크시트는 무엇인가요?", F(13, True, NAVY))
setc(ws, "B8", "‘하네스 첫걸음’ 실습을 직접 따라가며 기록하고, 나만의 첫 하네스를 설계해 보는 실습 노트입니다. 코딩 지식은 필요 없습니다.", F(10.5), align="left", wrap=True)
ws.merge_cells("B8:C8"); ws.row_dimensions[8].height = 34

rows = [
    ("① 실습 체크리스트", "‘제로톡’ 미션을 7단계로 따라가며, 각 단계의 완료 여부와 메모를 기록합니다."),
    ("② 하네스 설계 캔버스", "내 업무를 도메인·역할·스킬·산출물·패턴으로 쪼개, 나만의 하네스 요청문을 완성합니다."),
    ("③ 자기평가 루브릭", "오늘 배운 내용을 4개 항목으로 스스로 채점합니다. 점수는 자동 합산됩니다(70점 이상 통과)."),
]
r = 10
setc(ws, f"B{r}", "탭", F(10, True, "FFFFFF"), NAVY, "center", box)
setc(ws, f"C{r}", "이렇게 쓰세요", F(10, True, "FFFFFF"), NAVY, "left", box)
r += 1
for i, (a, b) in enumerate(rows):
    bg = "FFFFFF" if i % 2 == 0 else MISTBG
    setc(ws, f"B{r}", a, F(10.5, True, NAVY), bg, "left", box)
    setc(ws, f"C{r}", b, F(10), bg, "left", box, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
setc(ws, f"B{r}", "오늘의 미션", F(12, True, CORAL))
r += 1
setc(ws, f"B{r}", "신제품 음료 “제로톡” 런칭을 위한 미니 리서치+카피 팀을 만들어, 시장요약 1페이지와 광고 헤드라인 5개를 받는다.", F(10.5, False, INK), CORAL_BG, "left", box, wrap=True)
ws.merge_cells(f"B{r}:C{r}"); ws.row_dimensions[r].height = 30

r += 2
setc(ws, f"B{r}", "복붙 프롬프트", F(10, True, TEAL))
r += 1
setc(ws, f"B{r}", "하네스 구성해줘. 신제품 음료 런칭을 돕는 작은 팀이 필요해 — 시장/경쟁 상황을 빠르게 조사하는 리서처와, 그 조사를 바탕으로 광고 헤드라인을 뽑는 카피라이터. 조사 요약 1페이지와 헤드라인 5개를 결과로 줘.", F(10, False, "E2E8F0"), NAVY, "left", box, wrap=True)
ws.merge_cells(f"B{r}:C{r}"); ws.row_dimensions[r].height = 46

# ============================================================
# SHEET 2 — 실습 체크리스트
# ============================================================
ws = wb.create_sheet("실습 체크리스트")
ws.sheet_view.showGridLines = False
widths = [2.5, 6, 46, 34, 12, 26]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = w
title_block(ws, "실습 체크리스트 — “제로톡” 런칭 팀", 6)

hdr = 7
heads = ["단계", "할 일", "이렇게 되면 성공", "완료", "메모 (관찰한 것)"]
for i, h in enumerate(heads):
    setc(ws, f"{get_column_letter(i+2)}{hdr}", h, F(10, True, "FFFFFF"), NAVY, "center" if i in (0,3) else "left", box)
ws.row_dimensions[hdr].height = 22

steps = [
    ("1", "Claude Code 실행 → 플러그인 목록에서 harness 확인", "harness가 목록에 보임"),
    ("2", "에이전트 팀 기능이 켜져 있는지 확인", "팀 기능 ON"),
    ("3", "위 프롬프트를 그대로 입력", "리서처+카피라이터 2인 팀 구성"),
    ("4", ".claude/agents/ 와 .claude/skills/ 폴더를 눈으로 확인", "“누가”와 “어떻게”가 파일로 분리됨"),
    ("5", "실행 결과 받기", "시장요약 1p + 헤드라인 5개 확보"),
    ("6", "이 흐름을 “파이프라인(조사→카피)”으로 라벨링", "조사 → 카피 순서를 이해"),
    ("7", "“헤드라인을 더 위트있게”로 재요청", "같은 팀이 톤만 바꿔 재생성"),
]
r = hdr + 1
dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
ws.add_data_validation(dv)
for i, (n, task, exp) in enumerate(steps):
    bg = "FFFFFF" if i % 2 == 0 else MISTBG
    setc(ws, f"B{r}", n, F(11, True, CORAL), bg, "center", box)
    setc(ws, f"C{r}", task, F(10), bg, "left", box, wrap=True)
    setc(ws, f"D{r}", exp, F(10, False, TEALD), bg, "left", box, wrap=True)
    setc(ws, f"E{r}", "☐", F(13, False, CORAL), bg, "center", box)
    setc(ws, f"F{r}", "", F(10), bg, "left", box, wrap=True)
    dv.add(ws[f"E{r}"])
    ws.row_dimensions[r].height = 34
    r += 1

# progress counter (formula)
r += 1
setc(ws, f"C{r}", "완료한 단계 수 →", F(11, True, NAVY), None, "right")
setc(ws, f"D{r}", f'=COUNTIF(E{hdr+1}:E{hdr+len(steps)},"☑")&" / {len(steps)}"', F(11, True, CORAL), AMBER_BG, "center", box)
setc(ws, f"F{r}", "체크박스(완료 열)를 ☑ 로 바꾸면 자동 집계됩니다.", F(9, False, SLATE), None, "left", wrap=True)

r += 2
setc(ws, f"B{r}", "핵심 체험", F(11, True, CORAL))
r += 1
setc(ws, f"B{r}", "재요청 한 번으로 결과가 바뀝니다. 하네스는 고정물이 아니라 ‘쓸수록 진화하는 시스템’입니다.", F(10.5), CORAL_BG, "left", box, wrap=True)
ws.merge_cells(f"B{r}:F{r}"); ws.row_dimensions[r].height = 26

# ============================================================
# SHEET 3 — 하네스 설계 캔버스
# ============================================================
ws = wb.create_sheet("하네스 설계 캔버스")
ws.sheet_view.showGridLines = False
widths = [2.5, 22, 40, 40]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = w
title_block(ws, "하네스 설계 캔버스 — 나만의 팀 만들기", 4)

setc(ws, "B7", "내 업무를 쪼개 보고, 아래 빈칸을 채워 나만의 하네스 요청문을 완성하세요. 예시(제로톡)를 참고하세요.", F(10.5), align="left", wrap=True)
ws.merge_cells("B7:D7"); ws.row_dimensions[7].height = 28

hdr = 9
for col, h in zip("BCD", ["설계 항목", "예시 (제로톡)", "내 업무 (직접 작성)"]):
    setc(ws, f"{col}{hdr}", h, F(10, True, "FFFFFF"), NAVY, "left", box)
ws.row_dimensions[hdr].height = 22

canvas = [
    ("① 도메인 (무슨 일?)", "신제품 음료 “제로톡” 런칭"),
    ("② 필요한 역할 (누가?)", "리서처, 카피라이터"),
    ("③ 각 역할의 방법 (스킬)", "리서처: 시장·경쟁 요약 / 카피: 헤드라인 5개"),
    ("④ 산출물 (무엇을 받나?)", "시장요약 1페이지 + 헤드라인 5개"),
    ("⑤ 협업 패턴", "파이프라인 (조사 → 카피)"),
    ("⑥ 완성한 요청문", "하네스 구성해줘. 신제품 음료 런칭을 돕는 팀 — 리서처와 카피라이터. 시장요약 1p와 헤드라인 5개를 줘."),
]
r = hdr + 1
for i, (label, ex) in enumerate(canvas):
    bg = "FFFFFF" if i % 2 == 0 else MISTBG
    setc(ws, f"B{r}", label, F(10.5, True, NAVY), bg, "left", box, wrap=True)
    setc(ws, f"C{r}", ex, F(10, False, SLATE, it=True), bg, "left", box, wrap=True)
    setc(ws, f"D{r}", "", F(10), "FFFFFF", "left", box, wrap=True)
    ws.row_dimensions[r].height = 46 if i < 5 else 70
    r += 1

# highlight the "내 요청문" answer cell
ws[f"D{r-1}"].fill = fill(AMBER_BG)

r += 1
setc(ws, f"B{r}", "패턴 힌트", F(11, True, TEALD))
r += 1
hints = [
    ("파이프라인", "순차 의존 — 앞 결과가 다음 입력 (조사→카피)"),
    ("팬아웃/팬인", "여러 각도로 동시 조사 후 종합"),
    ("생성-검증", "만들고, 다른 에이전트가 검수"),
    ("감독자", "중앙 지휘자가 태스크 배분"),
]
setc(ws, f"B{r}", "패턴", F(9, True, "FFFFFF"), NAVY, "left", box)
setc(ws, f"C{r}", "이럴 때", F(9, True, "FFFFFF"), NAVY, "left", box)
ws.merge_cells(f"C{r}:D{r}")
r += 1
for i, (p, d) in enumerate(hints):
    bg = "FFFFFF" if i % 2 == 0 else MISTBG
    setc(ws, f"B{r}", p, F(10, True, TEALD), bg, "left", box)
    setc(ws, f"C{r}", d, F(9.5), bg, "left", box, wrap=True)
    ws.merge_cells(f"C{r}:D{r}")
    ws.row_dimensions[r].height = 20
    r += 1

# ============================================================
# SHEET 4 — 자기평가 루브릭
# ============================================================
ws = wb.create_sheet("자기평가 루브릭")
ws.sheet_view.showGridLines = False
widths = [2.5, 26, 44, 12, 14]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = w
title_block(ws, "자기평가 루브릭 (R-C1)", 5)

setc(ws, "B7", "각 항목을 스스로 1~3점으로 채점하세요. 1=미흡, 2=보통, 3=우수. 가중 점수는 자동 계산됩니다.", F(10.5), align="left", wrap=True)
ws.merge_cells("B7:E7"); ws.row_dimensions[7].height = 26

hdr = 9
for col, h in zip("BCDE", ["평가 항목", "우수(3점) 기준", "가중치", "내 점수(1-3)"]):
    setc(ws, f"{col}{hdr}", h, F(10, True, "FFFFFF"), NAVY, "center" if col in "DE" else "left", box)
ws.row_dimensions[hdr].height = 22

crit = [
    ("개념 이해", "오케스트레이터·에이전트·스킬 3층 구조를 광고팀 비유로 설명한다", 0.30),
    ("설치·실행", "독립적으로 설치·활성화·첫 하네스 실행을 완료한다", 0.30),
    ("요청문 품질", "도메인·역할·산출물 3요소를 한 문장에 명시한다", 0.25),
    ("검증 태도", "결과를 그대로 받지 않고 재요청으로 개선한다", 0.15),
]
r = hdr + 1
first = r
dv2 = DataValidation(type="whole", operator="between", formula1="1", formula2="3", allow_blank=True,
                     error="1에서 3 사이 정수를 입력하세요", errorTitle="점수 범위")
ws.add_data_validation(dv2)
for i, (name, desc, w) in enumerate(crit):
    bg = "FFFFFF" if i % 2 == 0 else MISTBG
    setc(ws, f"B{r}", name, F(10.5, True, NAVY), bg, "left", box, wrap=True)
    setc(ws, f"C{r}", desc, F(10), bg, "left", box, wrap=True)
    setc(ws, f"D{r}", w, F(10, False, INK), bg, "center", box)
    ws[f"D{r}"].number_format = "0%"
    setc(ws, f"E{r}", None, F(11, True, "0000FF"), AMBER_BG, "center", box)
    dv2.add(ws[f"E{r}"])
    ws.row_dimensions[r].height = 40
    r += 1
last = r - 1

# weighted total row
setc(ws, f"B{r}", "가중 합계 (100점 만점)", F(11, True, NAVY), MISTBG, "left", box)
ws.merge_cells(f"B{r}:D{r}")
# weighted score = sum(weight * score/3)*100
setc(ws, f"E{r}", f"=ROUND(SUMPRODUCT(D{first}:D{last},E{first}:E{last})/3*100,0)", F(12, True, INK), MISTBG, "center", box)
total_cell = f"E{r}"
ws.row_dimensions[r].height = 24
r += 1

# pass/fail
setc(ws, f"B{r}", "판정 (70점 이상 통과)", F(11, True, NAVY), None, "left", box)
ws.merge_cells(f"B{r}:D{r}")
setc(ws, f"E{r}", f'=IF(COUNT(E{first}:E{last})=0,"-",IF({total_cell}>=70,"통과 ✅","재도전 🔁"))', F(12, True, TEALD), TEAL_BG, "center", box)
ws.row_dimensions[r].height = 24

r += 2
setc(ws, f"B{r}", "점수 = Σ(가중치 × 내점수 ÷ 3) × 100. 파란색 셀에 점수를 입력하면 자동 계산됩니다.", F(9, False, SLATE), None, "left", wrap=True)
ws.merge_cells(f"B{r}:E{r}")

from openpyxl.worksheet.properties import PageSetupProperties
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 1
    sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sh.page_margins.left = sh.page_margins.right = 0.4
    sh.page_margins.top = sh.page_margins.bottom = 0.5

wb.save(OUT)
print("WORKSHEET WRITTEN")
